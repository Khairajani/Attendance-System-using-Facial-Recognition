# -*- coding: utf-8 -*-
"""
Created on Sun Mar 15 00:06:57 2020

@author: Himanshu Khairajani
"""

import datetime
import openpyxl
from face_recognition_final import __MAIN__

date=str(datetime.date.today())
time=str(datetime.datetime.now())

time=list(map(int,time.split(' ')[1].split(':')[0:2]))
date=list(map(int,date.split('-')))

Total_Students=['Akella','Amardeep','Anurag','Ayush','Bhuvi','Gazal','Himanshu','Jai','Kavyansh','Kuldeep','Loveleen','Meenakshi','Prahlad','Pratik','Rajiv','Saanika','Sanchita','Sindhu','Vivek Pal']
Months=['January','February','March','April','May','June','July','August','September','October','November','December']
sub=['Math','Physics','Chemistry']
def sheet(Time):
    if Time==0:
        return 0           #'Math'
    elif Time==10:
        return 1           #'Physics'
    elif Time==11:
        return 2           #'Chemistry'

def update():
    Path_of_image="..\test_all_small\4.jpg"
    Students=__MAIN__(Path_of_image)

    subject=sheet(time[0])

    wb = openpyxl.load_workbook(filename = 'attendance_update.xlsx')

    ws = wb.worksheets[subject]


    for i in range(2,len(Total_Students)+2):
        ws.cell(row=i, column=3).value = ws.cell(row=i, column=3).value + 1

    for student in Students:
        rn=Total_Students.index(student)  + 2
        ws.cell(row=rn, column=2).value = ws.cell(row=rn, column=2).value + 1
        ws.cell(row=rn, column=date[1]+3).value = ws.cell(row=rn ,column=date[1]+3).value + (str(date[2]) + ',')

    wb.save("attendance_update.xlsx")

if __name__ == "__main__":
    update()































# import xlrd
# loc = ("Math.xlsx")
# wb = xlrd.open_workbook(loc)
# sheet = wb.sheet_by_index(0)
# print(sheet.cell_value(0, 0) )
# print(sheet.nrows)
# print(sheet.ncols)



























# import pandas as pd

# Data=pd.read_excel(r'Math.xlsx', sheet_name='Math')
# print(Data)


# import tkinter as tk
# from tkinter import filedialog
# import pandas as pd

# root= tk.Tk()

# canvas1 = tk.Canvas(root, width = 300, height = 300, bg = 'lightsteelblue')
# canvas1.pack()

# def getExcel ():
#     global df

#     import_file_path = filedialog.askopenfilename()
#     df = pd.read_excel (import_file_path)
#     print (df)

# browseButton_Excel = tk.Button(text='Import Excel File', command=getExcel, bg='green', fg='white', font=('helvetica', 12, 'bold'))
# canvas1.create_window(150, 150, window=browseButton_Excel)

# root.mainloop()